#--------------------------------------------------------------------------
# Editor: Rupin Bapuji
# Edits: Adding export function so schedule can be converted to excel file
#--------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random as r


#ask a user for input
def promptUser():
    """Prompts the user using input() and returns 1,2,3 depending on daily, weekly, or monthly.
    """
    while True:
        periodSelection = input("How long is your study period?:\n1.Day\n2.Week\n3.Month\n")
        if periodSelection in ["1", "2", "3"]:
            # user subjects list
            subjectList = []

            while True:
                    try:
                        numberOfSubjects = int(input("How many subjects do you want to study?: "))
                    except ValueError:
                        print("\nInvalid selection, please try again\n")
                        continue

                    #checks for negative
                    if numberOfSubjects > 0 and numberOfSubjects < 12:
                        #adds to list
                        for i in range(numberOfSubjects):
                            subjectList.append(input(f"Enter a subject: {i+1}. "))
                        return (int(periodSelection), subjectList)
        else:
            print("\nInvalid selection, please try again\n")
    


#generate the schedule
def generateSchedule(n, subjectList):   
    """ n is the user selection
            morning: 7am-12pm
            afternoon: 12pm - 5pm
            evening: 5pm - 10pm
    """
    if n == 1:
        schedule = dayGenerate(subjectList)
        print(schedule)
        exportSchedule(schedule)
    elif n == 2:
        pass
    elif n == 3:
        pass
    

def dayGenerate(subjectList):
    daySchedule = []

    #asks for num of hours
    while True:
        try:
            userStudyTime = int(input("\nHow many hours do you want to study for? "))
        except ValueError:
            print("\nInvalid selection, please try again\n")
            continue
        if userStudyTime < 12 and userStudyTime > 0:
            break

    #ask for time of day
    while True:
        timeOfDay = input("When would you like to start studying?:\n1.Morning 7am\n2.Afternoon 12pm\n3.Evening 5pm\n")
        
        if timeOfDay in ["1", "2", "3"]:
            # defines starting times for calculation, random date given as we only care about time
            if timeOfDay == "1":
                startTime = datetime(year=2020,month=10,day=10,hour=7)
            elif timeOfDay == "2":
                startTime = datetime(year=2020,month=10,day=10,hour=12)
            elif timeOfDay == "3":    
                startTime = datetime(year=2020,month=10,day=10,hour=5)
            break

    # 1:3 for break to study ratio
    userStudyTimeMinutes = userStudyTime*60
    breakTime = userStudyTimeMinutes // (4 * len(subjectList))
    studyTime = breakTime * 3

    #convert to list
    subjectTime = startTime
    for subject in subjectList:
        #adds the time and subject to schedule then adds time used
        formattedSubjectBeforeTime = subjectTime.strftime("%-I:%M")
        formattedSubjectAfterTime = (subjectTime + timedelta(minutes=studyTime)).strftime("%-I:%M")
        daySchedule.append(f"{formattedSubjectBeforeTime}-{formattedSubjectAfterTime}: {subject}")
        subjectTime += timedelta(minutes=studyTime)

        #adds break to schedule then adds time used
        formattedSubjectBeforeTime = subjectTime.strftime("%-I:%M")
        formattedSubjectAfterTime = (subjectTime + timedelta(minutes=breakTime)).strftime("%-I:%M")
        daySchedule.append(f"{formattedSubjectBeforeTime}-{formattedSubjectAfterTime}: Break")
        subjectTime += timedelta(minutes=breakTime)

    return daySchedule


def exportSchedule(schedule):
    #random colors for subjects
    colors = ["00FF6E63", "00FFE563", "0098FC7C", "007CE1FC", \
        "007CBEFC", "007C8DFC", "00AF96FF", "00CC96FF", "00FF96EC"]

    thin = Side(border_style="thin", color="000000") #border style

    
    scheduleXLS = Workbook() #creating workbook

    sheet = scheduleXLS.active #chosen sheet
    sheet.title = "Schedule"

    filename = input("What would you like your schedule to be named? ") + ".xlsx"

    #TITLES
    a, b = sheet['A1'], sheet['B1']
    a.value, b.value = "Time", "Task"
    a.fill, b.fill = PatternFill("solid", "00FFCC00"), PatternFill("solid", "00FFCC00")
    a.border = Border(top = thin, bottom = thin, left = thin, right = thin)
    b.border = Border(top = thin, bottom = thin, left = thin, right = thin)
    a.alignment, b.alignment = Alignment(horizontal = "center"), Alignment(horizontal = "center")

    #index i for the schedule list, and y is for the Excel row to choose
    index, y = 0, 2

    #Safeguards to avoid infinite loop
    while y <= len(schedule) * 2 + 1 and index < len(schedule):
        merged = False #Bool to check if cell is a subject or a break

        schedElements = schedule[index].split() #splits into time range and task

        for c in range(1, 3): #only need columns 1 and 2

            if (y - 1) % 4 != 0: #checks if row is not a break, then merges cells
                sheet.merge_cells(start_row = y, start_column = c, end_row = y + 2, end_column = c)
                merged = True #bool changes

            cell = sheet.cell(row = y, column = c) #chooses cell to add value
            cell.border = Border(top = thin, bottom = thin, left = thin, right = thin)

            if c == 1: #c == 1 for time slots
                cell.value = schedElements[0].split("-")[0] #chooses time slot

                cell.fill = PatternFill("solid", "00DEDEDE") #grey
                cell.alignment = Alignment(horizontal = "center", vertical= "top")
            else:
                cell.value = schedElements[1] #task
                cell.alignment = Alignment(horizontal = "center", vertical = "center")
                

        if merged:
            y += 3 #row needs to jump 3
            cell.fill = PatternFill("solid", r.choice(colors)) #random color for subject
        else:
            y += 1
            cell.fill = PatternFill("solid", "007CFCF2") #light blue for breaks

        index += 1 #schedule elements incremented


            
    scheduleXLS.save(filename) #saved with chosen name

    print("Your schedule is now saved in this directory.\n")
    
    
if __name__ == "__main__":
    #stores choices in a set
    userChoices = promptUser()

    #calls the generate function based on their choices
    generateSchedule(userChoices[0], userChoices[1])
